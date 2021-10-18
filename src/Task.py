import fnmatch
import configparser
import glob
import tfxinit
import time
import logging
from enum import Enum
import codecs
import json
from shutil import copyfile
import os
import os.path
import uuid
import pandas as pd
import numpy as np
import shutil
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
from celery import Celery
import pygogo as gogo
from celery_redis_sentinel import register
from SupportFiles import SupportFiles
from Normalizer import Normalizer
from Web2nl import Web2nl
from GenerateGRXML import GenerateGRXML
from SpeechModel import SpeechModel
import sh
import csv
import zipfile
from collections import defaultdict
from multiprocessing import current_process
from celery.signals import worker_process_init
from config import data

@worker_process_init.connect
def fix_multiprocessing(**kwargs):
    try:
        current_process()._config
    except AttributeError:
        current_process()._config = {'semprefix': '/mp'}

log_format = '[%(levelname)s] %(asctime)s: [%(modeluuid)s] %(message)s'
date_format = '%Y-%m-%dT%H:%M:%S%z'
formatter = logging.Formatter(log_format, date_format)

class ModelType(Enum):
    CLASSIFIER = 1
    NORMALIZATION = 2

register()
sentinels = []
broker_url_host = data.get("BROKER_URL_HOST")
broker_url_port = data.get("BROKER_URL_PORT")
celery_result_backend_host = data.get("CELERY_RESULT_BACKEND_HOST")
celery_result_backend_port = data.get("CELERY_RESULT_BACKEND_PORT")
app = Celery('tasks')
for i in data.get("BROKER_TRANSPORT_OPTIONS_SENTINELS").split(","):
    sentinels.extend([(i,26379)])

class Config:
    BROKER_URL = f"redis-sentinel://{broker_url_host}:{broker_url_port}/0"
    BROKER_TRANSPORT_OPTIONS = {
        'sentinels': sentinels,
        'service_name': data.get("BROKER_TRANSPORT_OPTIONS_SERVICE_NAME"),
        'socket_timeout': 1.0,
    }

    CELERY_RESULT_BACKEND = f"redis-sentinel://{celery_result_backend_host}:{celery_result_backend_port}/0"
    CELERY_RESULT_BACKEND_TRANSPORT_OPTIONS = BROKER_TRANSPORT_OPTIONS
    CELERYD_PREFETCH_MULTIPLIER = 1
app.config_from_object(Config)

#basePath = "/Users/huzefa.siyamwala/Code/huzefa/orion"
# basePath = "/var/tellme/modelbuilder_worker"
basePath = data.get('MODELBUILDER_ROOT_DIRECTORY','/var/tellme/modelbuilder_worker')

workbench_location = "/var/tellme/workbench/workspace/DSG-huzefa-siyamwala/statistical-semantic-interpretter/SSI-1.0/"
#workbench_location = "/Users/huzefa.siyamwala/Code/dsg/workbench//SSI-1.0/"


def generate_workbench_config_file(workDirectory, modelConfigFilePath, trainWithCrossValidation, supportFileHandler):
    config = configparser.RawConfigParser()
    config.optionxform = str
    config.read("configs/dsg_workbench_default_config.cfg")
    config['configs']['output_folder'] = workDirectory + "/Results"
    config['configs']['data_file'] = workDirectory + "/FirstResponses_WCNormalized"

    if supportFileHandler.get_training_configs_property('experiment_type') == "grid_search_new":
        config['configs']['experiment_type'] = "grid_search_new"
    else:
        if trainWithCrossValidation:
            config['configs']['experiment_type'] = "train_with_CV"
        else:
            config['configs']['experiment_type'] = "train"
    # SVM Light location
    config['configs']['svm_light_dir'] = "/var/tellme/workbench/workspace/DSG-huzefa-siyamwala/statistical-semantic-interpretter/SSI-1.0/resources/svm_light"
    #config['configs']['svm_light_dir'] = "/Users/huzefa.siyamwala/.workbench/SSI-1.0/resources/svm_light"
    stopWordFileLocation = supportFileHandler.get_stopwords_fileName()
    if stopWordFileLocation is not None:
        config['configs']['stopwords_file'] = stopWordFileLocation

    stemmingExceptionFileLocation = supportFileHandler.get_stemming_exceptions_filename()
    if stemmingExceptionFileLocation is not None:
        config['configs']['stemming_exceptions_file'] = stemmingExceptionFileLocation

    contractionFileLocation = supportFileHandler.get_contractions_file_name()
    if contractionFileLocation is not None:
        config['configs']['word_expansion_file'] = contractionFileLocation
    noOfFolds = supportFileHandler.get_training_configs_property('validationSplit')
    if noOfFolds is not None:
        config['configs']['number_of_folds'] = noOfFolds
    noOfFoldsPerSigmoidTraining = supportFileHandler.get_training_configs_property('numOfEpochs')
    if noOfFoldsPerSigmoidTraining is not None:
        config['configs']['number_of_folds_Sgmd_Training'] = noOfFoldsPerSigmoidTraining
    additionalTrainingConfigs = supportFileHandler.get_additional_training_configs_property()
    for key in additionalTrainingConfigs:
        if key=="experiment_type":
            continue
        config['configs'][key] = str(additionalTrainingConfigs[key])
    config_file_location = workDirectory+ "/Config.cfg"
    with open(config_file_location, 'w') as configfile:
        config.write(configfile)
    with open(config_file_location, 'r') as fin:
        data = fin.read().splitlines(True)
    with open(config_file_location, 'w') as fout:
        fout.writelines(data[1:])


def getAccuracy(work_directory, logger, report_type="internal"):
    try:
        path = os.path.join(work_directory, 'Results/aggregate_metrics_' + report_type + '_cumulative.csv')
        df = pd.read_csv(path, skiprows=0)
        weighted_fscore = df.iloc[0]['Average']
        accuracy = df.iloc[1]['Average']
        logger.info("Weighted F_Score:" + str(weighted_fscore))
        logger.info("Accuracy:" + str(accuracy))
        return str(weighted_fscore), str(accuracy)
    except Exception:
        raise RuntimeError("Failed to find aggregate metrics files!!")

def getMisalignedIntents(work_directory, accuracy, logger, report_type="internal"):
    try:
        path = os.path.join(work_directory, 'Results/ConfusionMatrix_' + report_type + '*.csv')
        columns = ['actual_intent', 'predicted_intent', 'count']
        df_stats = pd.DataFrame(columns=columns)
        df_stats = df_stats.fillna(0)

        for fname in glob.glob(path):
            logger.info("Processing: " + fname)
            df = pd.read_csv(fname, skiprows=1)
            # Dropping off Actual Intent column
            df.drop(df.columns[0], inplace=True, axis=1)
            predictedIntentColumnName = list(df.columns.values)
            predictedIntentColumnName = predictedIntentColumnName[1:]
            for index, row in df.iterrows():
                for index_1, element in enumerate(df.ix[index, 1:]):
                    df_stats.loc[-1] = [df.ix[index, 0], predictedIntentColumnName[index_1], element]
                    df_stats.index = df_stats.index + 1  # shifting index
            df_stats = df_stats[df_stats.actual_intent != df_stats.predicted_intent]
            df_stats = df_stats[df_stats['count'] > 0]

        wb = Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = 'Weighted F_Score'
        ws.cell(row=1, column=2).value = accuracy[0]

        ws.cell(row=2, column=1).value = 'Accuracy'
        ws.cell(row=2, column=2).value = accuracy[1]

        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)

        ws['A2'].font = Font(bold=True)
        ws['B2'].font = Font(bold=True)

        if df_stats.empty:
            logger.info("No misaligned intents!!")
        else:
            path = os.path.join(work_directory, 'Results/ClassificationOutput_' + report_type + '*.csv')
            list_ = []
            for fname in glob.glob(path):
                df_ques = pd.read_csv(fname)
                df_ques = df_ques[df_ques['Correct Classification'] == False]
                df_ques = df_ques.drop(df_ques.columns[[0, 4, 5, 6, 7, 8, 9]], axis=1)
                list_.append(df_ques)
            frame = pd.concat(list_)
            frame = frame.groupby(['Original Intent', 'Classified Intent 1'], as_index=False)['Utterance'] \
                .agg(lambda col: ';'.join(str(x) for x in col))
            frame = frame.rename(columns={'Original Intent': 'actual_intent', 'Classified Intent 1': 'predicted_intent',
                                          'Utterance': 'misaligned_questions'})

            df_stats = df_stats.groupby(['actual_intent', 'predicted_intent'], as_index=False).sum()
            result = pd.merge(df_stats, frame, on=['actual_intent', 'predicted_intent'])
            result = result.sort_values(by=['count'], ascending=[False])
            #result = result.applymap(lambda x: x.encode('unicode_escape').
            #                         decode('utf-8') if isinstance(x, str) else x)
            for r in dataframe_to_rows(result, index=False, header=True):
                ws.append(r)

            for cell in ws[3]:
                cell.style = 'Pandas'
            # Here adding three extra rows to account for weighted F1 scroe row, accuracy row and header row
            cell_range = 'C4:C' + str(len(result.index) + 3)
            redFill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
            yellowFill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type='solid')
            grayFill = PatternFill(start_color='c0c0c0', end_color='c0c0c0', fill_type='solid')
            ws.conditional_formatting.add(cell_range,
                                          CellIsRule(operator='greaterThanOrEqual', formula=['3'], stopIfTrue=True,
                                                     fill=redFill))
            ws.conditional_formatting.add(cell_range,
                                          CellIsRule(operator='lessThanOrEqual', formula=['1'], stopIfTrue=True,
                                                     fill=grayFill))
            ws.conditional_formatting.add(cell_range,
                                          CellIsRule(operator='equal', formula=['2'], stopIfTrue=True, fill=yellowFill))

        wb.save(work_directory + "/model_stats_" + report_type + ".xlsx")
        if report_type == "internal":
            os.symlink(os.path.join(work_directory, "model_stats_internal.xlsx"), os.path.join(work_directory, "model_stats.xlsx"))
    except Exception:
        raise


def updateNormalizerConfigFile(workDirectory, modelConfigFilePath):
    dataNormalizerConfig = workDirectory + "/normalizer.cfg"
    # Load the configuration file
    config = configparser.ConfigParser()
    config.read("configs/StandardDataNormalizer-Config.cfg")
    cfgfile = open(dataNormalizerConfig, 'w')
    config.write(cfgfile)
    cfgfile.close()


def getWeb2NLConverterCommand(workDirectory, modelConfigFilePath):
    command = "java -Dfile.encoding=UTF-8 -cp  "
    command += workbench_location
    command += "SSI-1.0.jar:"
    command += workbench_location
    command += "lib/* com.tfs.SSI.ModelingWorkbenchModelGenerator "
    # if config.json is not provided, using default transformation configs
    if os.path.exists(os.path.join(workDirectory, "config.json")):
        command += os.path.join(workDirectory, "config.json")
    else:
        command += "supporting_files/config.json"
    command += " "
    command += os.path.join(workDirectory, "final_model")
    command += ">>" + workDirectory + "/training.log"
    return command


def getWeb2NLNormalizationModelCommand(workDirectory, modelConfigFilePath):
    command = "java -Dfile.encoding=UTF-8 -cp  "
    command += workbench_location
    command += "SSI-1.0.jar:"
    command += workbench_location
    command += "lib/* com.tfs.SSI.ModelingWorkbenchModelGenerator "
    # if config.json is not provided, using default transformation configs
    if os.path.exists(os.path.join(workDirectory, "normalization.json")):
        command += os.path.join(workDirectory, "normalization.json")
    else:
        command += "supporting_files/config.json"
    command += " "
    command += os.path.join(workDirectory, "final_model")
    command += ">>" + workDirectory + "/training.log"
    return command


def getModelTrainingCommand(workDirectory):
    command = "java -server -d64 -XX:+UseParallelGC -XX:+UseParallelOldGC "
    command += "-Xms4G -Xmx90G "
    command += "-Dfile.encoding=UTF-8 "
    command += "-Dlog4j.configuration="
    command += workbench_location
    command += "log4j.properties -cp "
    command += workbench_location
    command += "SSI-1.0.jar:"
    command += workbench_location
    command += "lib/* com.tfs.SSI.SSI_mod "
    command += os.path.join(workDirectory, "Config.cfg")
    command += ">>" + workDirectory + "/training.log"
    return command


def copyDefaultNormalizationModel(work_directory):
    destinationPath = work_directory + "/final_model/web2nl/"
    os.makedirs(os.path.dirname(destinationPath), exist_ok=True)
    copyfile("./supporting_files/defaultNormalization.model", destinationPath + "/final.model")
    return


def copyDefaultConfigFile(work_directory):
    destinationPath = work_directory
    os.makedirs(os.path.dirname(destinationPath), exist_ok=True)
    copyfile("./supporting_files/config.json", destinationPath + "/config.json")
    return


def writeWorkbenchLogsToLogFiles(work_directory, logger):
    if os.path.exists(os.path.join(work_directory, "Results")):
        for file in os.listdir(os.path.join(work_directory, "Results")):
            if fnmatch.fnmatch(file, 'Log_*'):
                logFile = os.path.join(os.path.join(work_directory, "Results"), file)
                with open(logFile) as f:
                    read_data = f.read()
                    logger.info(read_data)


def setupForLocalDevelopment():
    data["WEB2NL_URL"] = "http://stable.api.sv2.247-inc.net/v1/classifier/"
    data["WEB2NL_API_KEY"] = "YYjr2deysZmsKTv2"
    # data["MODEL_BUILDER_STORAGE"] = "/Users/huzefa.siyamwala/Code/huzefa/orion-web-service/test/upload/"
    data["MODEL_BUILDER_STORAGE"] = "/Users/huzefa.siyamwala/Code/huzefa/orion/upload/"
    data["MODELBUILDER_URL"] = "http://localhost:8081/v1/modelbuilder/"
    data["PYTHONIOENCODING"] = "utf-8"
    data["LOG_LOCATION"] = "../logs/modelbuilder-worker.log"
    data["SRILM_TOOLS_LOCATION"]= "/Users/huzefa.siyamwala/Code/huzefa/orion/supporting_files/srilm_binaries" \
                                     "/macosx/ngram-count"
    data["MICROSOFT_SDK_HOST"] = "fafr25.pool.sv2.247-inc.net"
    data["RECO_HOST"] = "rec04.p303.sv2.247-inc.net"



def updateConfigFileForModelType(modelUUID, type):
    with codecs.open(os.path.join(os.path.join(data['MODEL_BUILDER_STORAGE'], modelUUID), "config.json"), "r+",
                     encoding='utf-8') as json_data:
        configData = json.load(json_data)
        json_data.seek(0)
        configData["modelType"] = type.name.lower()
        json_data.write(json.dumps(configData))
        json_data.truncate()


def generateConfigFileForNormalization(modelUUID):
    with codecs.open(os.path.join(os.path.join(data['MODEL_BUILDER_STORAGE'], modelUUID), "config.json"), "r+",
                     encoding='utf-8') as json_data:
        configData = json.load(json_data)
        json_data.seek(0)
        configData["modelType"] = ModelType.NORMALIZATION.name.lower()
        with codecs.open(os.path.join(os.path.join(data['MODEL_BUILDER_STORAGE'], modelUUID),
                                      "normalization.json"),
                         "w+",
                         encoding='utf-8') as normalized_json:
            # Remove input match Transformation from normalization model
            transformList = configData["transformations"]
            transformList = [tup for tup in transformList if isInputMatch(tup)]
            configData["transformations"] = transformList

            # Remove post processing rules from normalization model
            if "postProcessingRules" in configData:
                del configData["postProcessingRules"]
            normalized_json.write(json.dumps(configData))
            normalized_json.truncate()


def isInputMatch(element):
    if type(element) is dict:
        for name in element:
            if element[name]["type"] == "input-match":
                return False
    return True


def executeSystemCommand(command):
    return_status = os.system(command)

    if return_status is not 0:
        raise RuntimeError("Failed while training the model with workbench")


def copyNormalizationModelToUUIDPath(normalization_uuid, work_directory):
    # move normalization model to normalization uuid path
    normalization_model_directory = os.path.join(os.path.join(data['MODEL_BUILDER_STORAGE'],
                                                              normalization_uuid), "final_model/web2nl/")
    os.makedirs(os.path.dirname(normalization_model_directory), exist_ok=True)
    copyfile(os.path.join(work_directory, "final_model/web2nl/final.model"), os.path.join(normalization_model_directory, "final.model"))
    return normalization_uuid


def addSpeechCapability(workDirectory, modelUUID, speechConfigs, isUnbundled, digitalHostedUrl, logger):
    logger.info("Generating GRXML files")
    executeSystemCommand("mkdir -p " + workDirectory + "/grxml")
    grxml_handler = GenerateGRXML(workDirectory, modelUUID, logger)
    grxml_handler.generateGrammerFiles()
    logger.info("GRXML files generated successfully")
    speechHandler = SpeechModel(workDirectory, modelUUID, isUnbundled, digitalHostedUrl, logger)
    # generateSLMModel if digitalHostedUrl is null else generateSpeechModel
    if digitalHostedUrl is None or digitalHostedUrl == "":
        logger.info("digitalHostedUrl is not set ; Building SLM Model only")
        speechHandler.generateSLMModel()
    else:
        speechHandler.generateSpeechModel()
    return

def combineSLMAndSSIModel(workDirectory, modelUUID, isUnbundled, digitalHostedUrl, logger):
    logger.info("Combining SLM model with SSI")
    speechHandler = SpeechModel(workDirectory, modelUUID, isUnbundled, digitalHostedUrl, logger)
    speechHandler.generateCombinedModel()
    return

def parseClassificationOutputs(work_directory, logger):
    dic = defaultdict(list)
    with open(work_directory + "/Results/sklearn/ClassificationOutput_combined.csv", mode= "r") as f:
        reader = csv.reader(f)
        next(reader)
        for row in reader:
            dic[(row[2],row[3])].append(row[1])

    #Filter all which are incorrectly predicted
    incorrectpredictons = { key:value for (key,value) in dic.items() if key[0] != key[1]}

    totalIncorrectPredictions = 0
    for key,value in incorrectpredictons.items():
        totalIncorrectPredictions+=len(value)
    totalCorrectPredictions = 0
    for key,value in dic.items():
        totalCorrectPredictions+=len(value)
    accuracy = (totalCorrectPredictions-totalIncorrectPredictions)/totalCorrectPredictions

    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1).value = 'Accuracy'
    ws.cell(row=1, column=2).value = accuracy

    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    row = ["actual_intent","predicted_intent","count","misaligned_questions"]
    ws.append(row)
    if totalIncorrectPredictions == 0:
        logger.info("No misaligned intents!!")
    else:
        incorrectpredictonssorted = sorted(incorrectpredictons, key=lambda k: len(incorrectpredictons[k]),
                                           reverse=True)
        for key in incorrectpredictonssorted:
            row = []
            row.append(str(key[0]))
            row.append(str(key[1]))
            row.append((len(incorrectpredictons[key])))
            row.append(str(incorrectpredictons[key]))
            ws.append(row)


        for cell in ws[2]:
            cell.style = 'Pandas'
        # Here adding two extra rows to account for accuracy row and header row
        cell_range = 'C3:C' + str(len(incorrectpredictons.keys()) + 2)
        redFill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
        yellowFill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type='solid')
        grayFill = PatternFill(start_color='c0c0c0', end_color='c0c0c0', fill_type='solid')
        ws.conditional_formatting.add(cell_range,
                                      CellIsRule(operator='greaterThanOrEqual', formula=['3'], stopIfTrue=True,
                                                 fill=redFill))
        ws.conditional_formatting.add(cell_range,
                                      CellIsRule(operator='lessThanOrEqual', formula=['1'], stopIfTrue=True,
                                                 fill=grayFill))
        ws.conditional_formatting.add(cell_range,
                                      CellIsRule(operator='equal', formula=['2'], stopIfTrue=True, fill=yellowFill))

    wb.save(work_directory + "/model_stats.xlsx")


@app.task
def generateModel(modelUUID, modelType, isUnbundled, digitalHostedUrl=None, trainingFile=None, tfClient=None, modelTechnology=None, modelName=None, vectorizerVersion=None):
    build_start_time = time.time()
    start_time = time.time()
    #setupForLocalDevelopment()
    modelStoragePath = data.get("MODEL_BUILDER_STORAGE").split(",")[0]
    data["MODEL_BUILDER_STORAGE"] = modelStoragePath

    # Setting up logger with modelUUID as prefix
    ModelLogger = gogo.Gogo(__name__, low_hdlr=gogo.handlers.file_hdlr(data.get("LOG_LOCATION")),
                            low_level='info',
                            low_formatter=formatter, high_formatter=formatter)
    Modellogger = ModelLogger.get_logger(modeluuid=modelUUID)
    Modellogger.debug('Starting to create a model')

    # Setting up work directory for model building
    workDirectory = os.path.join(data['MODEL_BUILDER_STORAGE'], modelUUID)
    Modellogger.info("Setting work directory to %s", workDirectory)

    os.chdir(basePath)
    modelTrainingConfigFilePath = None
    postProcessingRules = None
    try:
        # Parsing config.json file
        if os.path.exists(workDirectory + "/config.json"):
            supportFileHandler = SupportFiles(os.path.join(workDirectory, "config.json"), workDirectory, Modellogger)
            supportFileHandler.parse_config_file()
            with codecs.open(os.path.join(os.path.join(data['MODEL_BUILDER_STORAGE'], modelUUID), "config.json"), "r+",
                     encoding='utf-8') as json_data:
                configData = json.load(json_data)
                if "postProcessingRules" in configData:
                    postProcessingRules = configData["postProcessingRules"]
        else:
            supportFileHandler = SupportFiles(None, workDirectory, Modellogger)

        if modelType == "speech":
            Modellogger.info("Creating speech model ")
            updateStatus(workDirectory, "Creating speech model")
            speechConfigs = supportFileHandler.get_speech_configs_section()
            # call corpus creation func if training file is sent

            # If trainingFile is sent then normalize the dataset
            if trainingFile is not None:
                Modellogger.info("Normalizing Speech Training File")
                normalizer_handler = Normalizer(None, modelUUID, Modellogger, None)
                normalizer_handler.normalizeSpeech(trainingFile) 
                                     
            addSpeechCapability(workDirectory, modelUUID, speechConfigs, isUnbundled, digitalHostedUrl, Modellogger)
            os.chdir(workDirectory)
            generateTrainingZipWithSpeech(workDirectory, Modellogger)
            os.chdir(basePath)
            updateStatus(workDirectory, "Successfully created speech model")
            writeWorkbenchLogsToLogFiles(workDirectory, Modellogger)
            Modellogger.info("Successfully created speech model")
            end_time = time.time()
            Modellogger.info("Adding speech took %s seconds", str(end_time - start_time))
            return
        elif modelType == "combine":
            combineSLMAndSSIModel(workDirectory, modelUUID, isUnbundled, digitalHostedUrl, Modellogger)
            os.chdir(basePath)
            updateStatus(workDirectory, "Successfully combined SLM model with SSI model")
            writeWorkbenchLogsToLogFiles(workDirectory, Modellogger)
            Modellogger.info("Successfully combined SLM model with SSI model")
            end_time = time.time()
            Modellogger.info("Combining SLM and SSI models took %s seconds", str(end_time - start_time))
            return
        else:
            if (modelTechnology != None and modelTechnology != 'n-gram'):
                # Instantiate Normalizer instance for current task
                normalization_uuid = str(uuid.uuid4())
                intent=supportFileHandler.get_default_intent()
                tfxinit.tfx_normalization(modelUUID, Modellogger,intent,normalization_uuid,tfClient,modelTechnology,modelName,vectorizerVersion,postProcessingRules)
            else:
                ''' Normalization Block'''
                # if normalization model without config needs to be created, just copying the default model
                Modellogger.info('Creating normalization model')
                if not os.path.exists(workDirectory + "/config.json"):
                    Modellogger.info('No configuration uploaded for creating normalization model, using default one')
                    copyDefaultNormalizationModel(workDirectory)
                    copyDefaultConfigFile(workDirectory)
                    updateStatus(workDirectory, "Normalization model Created Successfully")
                    Modellogger.info("Normalization model Created Successfully")
                else:
                    Modellogger.info('Using user uploaded configs for creating normalization model')
                    generateConfigFileForNormalization(modelUUID)
                    os.system(getWeb2NLNormalizationModelCommand(workDirectory, None))
                    updateStatus(workDirectory, "Normalization model Created Successfully")
                    Modellogger.info("Normalization model Created Successfully")

                if modelType == "normalization":
                    # We are done here, returning
                    return

                Modellogger.info('Starting to create a classifier model')
                # Instantiate Normalizer instance for current task
                normalization_uuid = str(uuid.uuid4())
                Modellogger.info("Creating Normalization model with uuid: %s", normalization_uuid)
                copyNormalizationModelToUUIDPath(normalization_uuid,workDirectory)
                normalizer_handler = Normalizer(normalization_uuid, modelUUID, Modellogger,
                                                supportFileHandler.get_default_intent())
                
                # Normalizing input data
                updateStatus(workDirectory, "Normalizing data via web2nl")
                normalizer_handler.normalize_text()
                end_time = time.time()
                Modellogger.info("Normalization took %s seconds", str(end_time - start_time)) 
                Modellogger.info("Normalization took %s seconds", str(end_time - start_time)) 
                Modellogger.info("Normalization took %s seconds", str(end_time - start_time)) 

                ''' SVM Model Training block '''
                start_time = time.time()
                updateConfigFileForModelType(modelUUID, ModelType.CLASSIFIER)

                generate_workbench_config_file(workDirectory, modelTrainingConfigFilePath, False, supportFileHandler)
                updateStatus(workDirectory, "Normalization completed, starting to build model via workbench...")
                Modellogger.info("Starting to build SVM model")

                deleteResultFolder(os.path.join(workDirectory, "Results"), Modellogger)

                executeSystemCommand(getModelTrainingCommand(workDirectory))

                executeSystemCommand("mkdir -p " + workDirectory + "/final_model")
                if supportFileHandler.get_training_configs_property('experiment_type') == "grid_search_new":
                    copyfile(os.path.join(workDirectory, "Results/final.model"), os.path.join(workDirectory, "final_model/final.model"))
                else:
                    copyfile(os.path.join(workDirectory, "Results/final.model"), os.path.join(workDirectory, "final_model/final.model"))
                updateStatus(workDirectory, "Model Building completed, converting it to web2NL format...")
                executeSystemCommand(getWeb2NLConverterCommand(workDirectory, modelTrainingConfigFilePath))
                updateStatus(workDirectory, "Training Completed. Calculating Training loss!!")
                Modellogger.info("Training Completed. Calculating Training loss!!")
                end_time = time.time()
                Modellogger.info("Total execution time for training: %s seconds", str(end_time - start_time))

                #With grid_search_new cross validation is done before training so no need of training
                if supportFileHandler.get_training_configs_property('experiment_type') == "grid_search_new":
                    parseClassificationOutputs(workDirectory,Modellogger)
                else:
                    ''' Cross Validation block  (default)'''
                    start_time = time.time()
                    generate_workbench_config_file(workDirectory, modelTrainingConfigFilePath, True, supportFileHandler)
                    Modellogger.info("Starting to create model with cross validation")
                    executeSystemCommand(getModelTrainingCommand(workDirectory))
                    accuracy = getAccuracy(workDirectory, Modellogger)
                    getMisalignedIntents(workDirectory, accuracy, Modellogger)
                    accuracy = getAccuracy(workDirectory, Modellogger, 'external')
                    getMisalignedIntents(workDirectory, accuracy, Modellogger, 'external')
                    end_time = time.time()
                    Modellogger.info("Total execution time for calculating accuracy: %s seconds", str(end_time - start_time))

                os.chdir(workDirectory)

                generateTrainingZip(workDirectory, Modellogger)

                os.chdir(basePath)

                # Validating model
                if Web2nl(modelUUID, Modellogger).is_model_valid() is False:
                    Modellogger.error("Model generated is invalid")
                    updateStatus(workDirectory, "Model generated is invalid")
                    raise RuntimeError("Model generated is invalid!!")

                if supportFileHandler.get_speech_configs_section() is not None:
                    updateStatus(workDirectory, "Digital model created successfully, adding speech capability ...")
                    Modellogger.info("Digital model created successfully, adding speech capability")
                    speechConfigs = supportFileHandler.get_speech_configs_section()
                    addSpeechCapability(workDirectory, modelUUID, speechConfigs, isUnbundled, digitalHostedUrl, Modellogger)
                    os.chdir(workDirectory)
                    generateTrainingZipWithSpeech(workDirectory, Modellogger)
                    os.chdir(basePath)

                updateStatus(workDirectory, "Model Created Successfully")
                writeWorkbenchLogsToLogFiles(workDirectory, Modellogger)
                Modellogger.info("Model Created Successfully")

    except UnicodeDecodeError as error:
        writeWorkbenchLogsToLogFiles(workDirectory, Modellogger)
        Modellogger.exception("Failed while training model: ", error)
        updateStatus(workDirectory, "Failed while decoding training data in utf-8 format")
    except Exception as e:
        writeWorkbenchLogsToLogFiles(workDirectory, Modellogger)
        Modellogger.exception("Failed while training model: ")
        updateStatus(workDirectory, "Model building failure: " + str(e))
        updateStatus(workDirectory, "Model building failure")
    total_time = (time.time() - build_start_time)
    with open(workDirectory+"/Results/aggregate_metrics_cumulative.csv","a") as fd:
        fd.write("\n")
        fd.write(f"Model Building Time: {total_time} sec")
    
    file_paths = tfxinit.get_all_file_paths(f'{workDirectory}/Results')
    with zipfile.ZipFile(f'{workDirectory}/trainingOutputs.zip','w') as zip:
        # writing each file one by one
        for file in file_paths:
            zip.write(file)
    

def generateTrainingZipWithSpeech(work_directory, logger):
    logger.info("Generating trainingOutputs.zip files from the Results, grxml, and Inter folders")
    trainingOutputsZipPath = os.path.join(work_directory, "trainingOutputs.zip")
    resultFolderPath = os.path.join(work_directory, "Results")
    sh.zip("-r", trainingOutputsZipPath, "./Results",
                   "./grxml","./Inter")
    deleteResultFolder(resultFolderPath, logger)
    
def generateTrainingZip(work_directory, logger):
    logger.info("Generating trainingOutputs.zip files from the Results folder")
    trainingOutputsZipPath = os.path.join(work_directory, "trainingOutputs.zip")
    sh.zip("-r", trainingOutputsZipPath, "./Results")

def updateStatus(work_directory, status):
    with open(work_directory + "/status", 'a') as out:
        out.write(status + '\n')

def deleteResultFolder(files_to_delete, logger):
    logger.info("Deleting Results directory")
    os.system("rm -rf " + files_to_delete)

if __name__ == '__main__':
    generateModel("huzefa-speech", "digital", True, "")
    # generateModel("ddf72f29-a32e-4365-8b72-08c8eb23ef6f")
    #WorkDir = "/Users/abhinav.gupta/Downloads/"
    #accuracy = getAccuracy(WorkDir, "", report_type="external")
    #getMisalignedIntents(WorkDir, accuracy, "", report_type="external")
